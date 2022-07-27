from urllib import response
import requests
from bs4 import BeautifulSoup
import call_useragent
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

time_path = 'time.txt'
time = open(time_path, 'r')

# 開始行數，行數固定+10
row_set = time.read()
row_set = int(row_set)
time = open(time_path, 'w')

for page_num in range(1, 26):
    row = row_set
    time_save = row
    page = (row_set/10)+1-0.6
    print("本次開始行數:")
    print(row_set)
    print("本次須擷取頁數:")
    print(page_num)

    n = input("輸入 END 取消迴圈 / 繼續執行直接 Enter:")
    if n == "END":
        break

    html_page = page_num*10

    print(html_page)
    
    path = 'html.txt'
    f = open(path, 'r', encoding="utf-8")
    url = "https://scholar.google.com.tw/scholar?start="+str(html_page)+"&q=enlitech+quantum+efficiency&hl=zh-TW&as_sdt=0,5&as_ylo=2020&as_yhi=2020"# 从 &q 后面开始修改网址

    headers = {'User-Agent': call_useragent.random()}

    response = requests.get(url, headers=headers)


    wb = load_workbook('new.xlsx')
    sheet = wb['QE']

    title_column = 3 # 標題行數
    rider_column = 6 # 作者行數
    herf_column = 8 # 連結行數
    name_column = 4 # 期刊名稱行數
    date_column = 2 # 年份行數

    c = sheet.cell(column=3, row=5)
    d = sheet.cell(column=7, row=5)

    soup = BeautifulSoup(response.text, "html.parser")

    title = soup.select('#gs_res_ccl_mid > div > div > h3 > a')
    rider = soup.select('#gs_res_ccl_mid > div > div.gs_ri > div.gs_a')

    for item in title:
        title_save = sheet.cell(row= row, column=title_column)
        title_value = item.get_text()
        title_save.value = title_value

        herf_save = sheet.cell(row= row, column=herf_column)
        herf_value = item.get('href')
        herf_save.value = herf_value

        row = row + 1

    row = row_set

    for item in rider:
        rider_save = sheet.cell(row= row, column=rider_column) #設定行數
        name_save = sheet.cell(row= row, column=name_column)
        date_save = sheet.cell(row= row, column=date_column)

        string_item = item.get_text() #字串切割 "作者 - paper, date - link"
        rider_value = string_item.split("-")
        if "," in rider_value[0]:
            rider_save.value = rider_value[0] #"作者"
        else:
            i = 1
            while i <= 8:
                error_save = sheet.cell(row=row, column=i)
                error_save.fill = PatternFill(fill_type="solid", fgColor="D45C1A") #遇到特殊格式全行標記
                i=i+1
        if 1 in range(len(rider_value)):
            name_value = rider_value[1].split(",")
            name_valueer = name_value[0]
            name_valueer.replace(u'\xa0', '') #為了判斷需拿掉 \xa0 格式
        else:
            next

        if 1 in range(len(name_value)): #判斷是否需要人工補齊
            date_value = name_value[1]
            date_save.value = date_value
        else:
            date_value = "None"
            date_save.fill = PatternFill(fill_type="solid", fgColor="EA9E16")
            date_save.value = date_value

        if '…' in name_valueer: #判斷是否需要人工補齊
            name_save.fill = PatternFill(fill_type="solid", fgColor="EA9E16") 
            name_save.value = name_valueer
        else:
            name_save.value = name_valueer

        row = row + 1 

    wb.save("new.xlsx")

    print("完成")
    f.close()
    row_set = row

    print("已完成行數:")
    print(row_set)
    print("已完成頁數:")
    print(page)
    time.write(format(time_save))

time.close()
print("已完成行數:")
print(row_set-1)
print("已完成頁數:")
print(page_num)
